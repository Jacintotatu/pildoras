import pandas as pd
import calendar
from datetime import datetime
import random
import os
import openpyxl

def pedir_datos():
    anio = int(input("Introduce el año: "))
    mes = int(input("Introduce el mes (1-12): "))
    cantidad_total = int(input("Introduce la cantidad total (€): "))
    cantidad_total = round(cantidad_total / 10) * 10  # Redondear al múltiplo de 10 más cercano
    dias_festivos_str = input("¿Qué días fueron festivos? (ej: 1,6,12): ")
    dias_con_valores = int(input("¿Cuántos días tendrán valores (sin contar domingos ni festivos)?: "))

    dias_festivos = [int(d.strip()) for d in dias_festivos_str.split(",") if d.strip().isdigit()]
    return anio, mes, cantidad_total, dias_con_valores, dias_festivos

def obtener_dias_laborables(anio, mes, dias_festivos):
    _, dias_en_mes = calendar.monthrange(anio, mes)
    dias = [
        datetime(anio, mes, d)
        for d in range(1, dias_en_mes + 1)
        if datetime(anio, mes, d).weekday() != 6 and d not in dias_festivos
    ]
    return dias

def repartir_cantidades(dias, total, dias_con_valores):
    dias_seleccionados = random.sample(dias, dias_con_valores)
    dias_seleccionados.sort()

    datos = []
    dobles = []
    if dias_con_valores < 19:
        dobles = random.sample(dias_seleccionados, random.randint(4, min(8, len(dias_seleccionados))))

    cantidades = []
    for dia in dias_seleccionados:
        if dia in dobles:
            cantidades.append([dia, random.choice(range(50, 301, 10)), False])
            cantidades.append([dia, random.choice(range(50, 301, 10)), True])
        else:
            cantidades.append([dia, random.choice(range(50, 301, 10)), True])

    suma_actual = sum(c[1] for c in cantidades)
    diferencia = total - suma_actual

    # Ajuste con múltiplos de 10
    intentos = 0
    while diferencia != 0 and intentos < 10000:
        i = random.randint(0, len(cantidades) - 1)
        actual = cantidades[i][1]
        nuevo = actual + (10 if diferencia > 0 else -10)
        if 50 <= nuevo <= 300:
            cantidades[i][1] = nuevo
            diferencia += -10 if diferencia > 0 else 10
        intentos += 1

    total_final = sum(c[1] for c in cantidades)
    if total_final != total:
        raise ValueError(f"No se pudo ajustar el total exacto. Total esperado: {total}, obtenido: {total_final}")

    return cantidades

import os
import pandas as pd
from openpyxl import load_workbook
from openpyxl.styles import numbers
import calendar

def generar_excel(datos, anio, mes, total, archivo="facturacion.xlsx"):
    filas = []
    contador_factura = 1

    for fecha, cantidad, mostrar_total in datos:
        fila = [
            contador_factura if mostrar_total else "",  # FACTURA SIMPLIFICADA
            cantidad,                                   # CANTIDADES DIARIAS (número real)
            cantidad if mostrar_total else "",          # TOTAL (solo si es última línea del día)
            "",                                          # Columna en blanco
            fecha.strftime("%d/%m/%Y")                  # FECHA
        ]
        filas.append(fila)
        if mostrar_total:
            contador_factura += 1

    # Fila en blanco + resumen final
    filas.append(["", "", "", "", ""])
    nombre_mes = calendar.month_name[mes].capitalize()
    filas.append(["", f"Total {nombre_mes}", total, "", ""])

    # Crear DataFrame
    columnas = ["FACTURA SIMPLIFICADA", "CANTIDADES DIARIAS", "TOTAL", "", "FECHA"]
    df = pd.DataFrame(filas, columns=columnas)

    # Añadir al archivo existente si ya existe
    if os.path.exists(archivo):
        df_existente = pd.read_excel(archivo)
        df = pd.concat([df_existente, df], ignore_index=True)

    # Guardar sin formato primero
    df.to_excel(archivo, index=False, engine='openpyxl')

    # Aplicar formato de moneda a las columnas necesarias
    wb = load_workbook(archivo)
    ws = wb.active

    for row in ws.iter_rows(min_row=2, max_row=ws.max_row):
        for col_idx in [2, 3]:  # Columnas B y C (CANTIDADES DIARIAS y TOTAL)
            cell = row[col_idx - 1]
            if isinstance(cell.value, (int, float)):
                cell.number_format = '€#,##0'  # Formato numérico con símbolo de euro

    wb.save(archivo)
    print(f"✅ Archivo guardado correctamente con formato de moneda: {archivo}")



# -------------------------------
# EJECUCIÓN
# -------------------------------
if __name__ == "__main__":
    anio, mes, total, dias_con_valores, festivos = pedir_datos()
    dias_validos = obtener_dias_laborables(anio, mes, festivos)
    datos = repartir_cantidades(dias_validos, total, dias_con_valores)
    generar_excel(datos, anio, mes, total)
