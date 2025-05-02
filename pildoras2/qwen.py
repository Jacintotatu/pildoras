import openpyxl
from openpyxl import Workbook
from datetime import datetime, date, timedelta
import calendar
import random
import os

def obtener_datos():
    año = int(input("Introduce el año: "))
    mes = int(input("Introduce el mes (1-12): "))
    cantidad_total = float(input("Introduce la cantidad total a repartir: "))
    dias_fiesta = int(input("Introduce la cantidad de días de fiesta (sin contar domingos): "))
    dias_reciben_cifras = int(input(f"¿Cuántos días del mes {mes}/{año} van a recibir cifras (sin contar domingos)? "))

    return año, mes, cantidad_total, dias_fiesta, dias_reciben_cifras

def es_dia_laborable(fecha):
    # Comprobamos si es domingo o no
    if fecha.weekday() == 6:
        return False
    return True

def generar_fechas_laborables(año, mes, dias_reciben_cifras, dias_fiesta):
    fechas = []
    festivos = []

    print(f"Introduce las fechas de los {dias_fiesta} días de fiesta (formato DD/MM/YYYY):")
    for i in range(dias_fiesta):
        while True:
            try:
                fecha_str = input(f"Día de fiesta {i+1}: ")
                fecha_festivo = datetime.strptime(fecha_str, "%d/%m/%Y").date()
                if fecha_festivo.month != mes or fecha_festivo.year != año:
                    print("La fecha debe pertenecer al mes y año especificados. Vuelve a intentarlo.")
                    continue
                festivos.append(fecha_festivo)
                break
            except ValueError:
                print("Formato incorrecto. Por favor introduce la fecha en formato DD/MM/YYYY")

    # Generar todas las fechas del mes
    _, num_dias = calendar.monthrange(año, mes)
    for dia in range(1, num_dias + 1):
        fecha = date(año, mes, dia)
        if es_dia_laborable(fecha) and fecha not in festivos:
            fechas.append(fecha)

    # Seleccionar solo los días que recibirán cifras
    if dias_reciben_cifras > len(fechas):
        print(f"Solo hay {len(fechas)} días laborables disponibles (después de restar domingos y festivos).")
        dias_reciben_cifras = len(fechas)

    fechas_seleccionadas = sorted(random.sample(fechas, dias_reciben_cifras))
    return fechas_seleccionadas

def redondear_multiplo_10(numero):
    return round(numero / 10) * 10

def distribuir_cantidad(cantidad_total, fechas_seleccionadas):
    datos = []
    cantidad_restante = cantidad_total
    dias_con_dos_cantidades = 0

    if len(fechas_seleccionadas) < 19:
        dias_con_dos_cantidades = random.randint(4, 8)

    dias_una_cantidad = len(fechas_seleccionadas) - dias_con_dos_cantidades
    cantidades_diarias = []

    # Primero generamos días con una sola cantidad (siempre positiva y <= 350€)
    for _ in range(dias_una_cantidad):
        cantidad = redondear_multiplo_10(random.uniform(50, min(350, cantidad_total)))
        cantidad = max(50, min(350, cantidad))  # Asegurar rango estricto
        cantidades_diarias.append([int(cantidad)])
        cantidad_restante -= cantidad

    # Luego generamos días con dos cantidades (ambas <= 350€, sumadas <= 350€)
    for _ in range(dias_con_dos_cantidades):
        cantidad1 = redondear_multiplo_10(random.uniform(50, 300))  # Menor que 350
        cantidad2 = redondear_multiplo_10(random.uniform(50, 350 - cantidad1))  # Para que la suma no supere 350€
        cantidad1 = max(50, min(350, cantidad1))
        cantidad2 = max(50, min(350 - cantidad1, cantidad2))
        cantidades_diarias.append([int(cantidad1), int(cantidad2)])
        cantidad_restante -= (cantidad1 + cantidad2)

    # Ajustamos para que sume exactamente la cantidad total
    diferencia = round(cantidad_total - (cantidad_total - cantidad_restante), 2)
    if abs(diferencia) > 0.01:
        idx = random.randint(0, len(cantidades_diarias)-1)
        if len(cantidades_diarias[idx]) == 1:
            nuevo_valor = max(50, min(350, cantidades_diarias[idx][0] + int(diferencia)))
            cantidades_diarias[idx][0] = int(redondear_multiplo_10(nuevo_valor))
        else:
            ajuste = redondear_multiplo_10(diferencia / 2)
            nuevo_valor1 = max(50, min(350, cantidades_diarias[idx][0] + ajuste))
            nuevo_valor2 = max(50, min(350, 350 - nuevo_valor1, cantidades_diarias[idx][1] + ajuste))
            cantidades_diarias[idx][0] = int(redondear_multiplo_10(nuevo_valor1))
            cantidades_diarias[idx][1] = int(redondear_multiplo_10(nuevo_valor2))

    # Asociamos las cantidades con las fechas
    factura_num = 1
    for i, fecha in enumerate(fechas_seleccionadas):
        facturas = []
        for j, cantidad in enumerate(cantidades_diarias[i]):
            if j == 0:
                facturas.append({
                    'factura_num': factura_num,
                    'cantidad': cantidad,
                    'total': None if len(cantidades_diarias[i]) > 1 else int(cantidades_diarias[i][0]),
                    'fecha': fecha.strftime("%d/%m/%Y")
                })
                factura_num += 1
            else:
                facturas.append({
                    'factura_num': None,
                    'cantidad': cantidad,
                    'total': int(round(sum(cantidades_diarias[i]))),
                    'fecha': fecha.strftime("%d/%m/%Y")  # La fecha se repite en ambas filas
                })
        datos.extend(facturas)

    return datos

def nombre_mes_en_espanol(mes):
    meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
             "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    return meses[mes-1]

def crear_o_actualizar_excel(datos, año, mes, cantidad_total):
    nombre_archivo = f"Facturacion_{año}.xlsx"
    mes_nombre = nombre_mes_en_espanol(mes)

    # Crear libro si no existe
    if not os.path.exists(nombre_archivo):
        wb = Workbook()
        ws = wb.active
        ws.title = mes_nombre
    else:
        from openpyxl import load_workbook
        wb = load_workbook(nombre_archivo)
        if mes_nombre in wb.sheetnames:
            print(f"Ya existe un mes {mes_nombre} en el archivo. ¿Quieres sobrescribirlo?")
            respuesta = input("Escribe 's' para sí o cualquier otra tecla para no: ").lower()
            if respuesta == 's':
                del wb[mes_nombre]
                ws = wb.create_sheet(mes_nombre)
            else:
                mes_nombre += f"_{año}_{mes}"
                ws = wb.create_sheet(mes_nombre)
        else:
            ws = wb.create_sheet(mes_nombre)

    # Escribir encabezados
    ws['A1'] = "FACTURA SIMPLIFICADA"
    ws['B1'] = "CANTIDADES DIARIAS"
    ws['C1'] = "TOTAL"
    ws['D1'] = ""
    ws['E1'] = "FECHA"

    # Escribir datos
    fila = 2
    for registro in datos:
        if registro['factura_num'] is not None:
            ws[f'A{fila}'] = registro['factura_num']
        if registro['cantidad'] is not None:
            ws[f'B{fila}'] = f"{registro['cantidad']}€"
        if registro['total'] is not None:
            ws[f'C{fila}'] = f"{registro['total']}€"
        if registro['fecha'] is not None:
            ws[f'E{fila}'] = registro['fecha']
        fila += 1

    # Fila en blanco
    fila += 1

    # Fila de total
    ws[f'B{fila}'] = f"Total {mes_nombre}"
    ws[f'C{fila}'] = f"{int(cantidad_total)}€"

    # Ajustar anchos de columna
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 20

    wb.save(nombre_archivo)
    print(f"\nDatos guardados en {nombre_archivo}, hoja '{mes_nombre}'.")

def main():
    while True:
        print("\n=== Introducir datos para nuevo mes ===")
        año, mes, cantidad_total, dias_fiesta, dias_reciben_cifras = obtener_datos()

        print(f"\nGenerando fechas laborables para {año}-{mes}...")
        fechas_seleccionadas = generar_fechas_laborables(año, mes, dias_reciben_cifras, dias_fiesta)

        print(f"\nDistribuyendo {cantidad_total}€ entre {len(fechas_seleccionadas)} días laborables...")
        datos = distribuir_cantidad(cantidad_total, fechas_seleccionadas)

        print("\nGenerando archivo Excel...")
        crear_o_actualizar_excel(datos, año, mes, cantidad_total)

        continuar = input("\n¿Quieres introducir otro mes? (s/n): ").lower()
        if continuar != 's':
            break

    print("Programa finalizado.")

if __name__ == "__main__":
    main()
