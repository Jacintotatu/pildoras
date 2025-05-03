import dearpygui.dearpygui as dpg
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from datetime import date
import calendar
import random
import os 

def nombre_mes_en_espanol(mes):
    """Devuelve el nombre del mes en espa√±ol."""
    meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
            "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    return meses[mes - 1]

def generar_fechas_laborables(a√±o, mes, dias_reciben_cifras, dias_fiesta, festivos):
    """Genera las fechas laborables del mes, excluyendo domingos y festivos."""
    fechas = []

    # Validar los d√≠as festivos
    festivos_validos = []
    _, num_dias = calendar.monthrange(a√±o, mes)
    for dia in festivos:
        if 1 <= dia <= num_dias:
            fecha_festivo = date(a√±o, mes, dia)
            if fecha_festivo.weekday() != 6:  # Excluir domingos
                festivos_validos.append(fecha_festivo)

    # Generar todas las fechas del mes
    for dia in range(1, num_dias + 1):
        fecha = date(a√±o, mes, dia)
        if fecha.weekday() != 6 and fecha not in festivos_validos:
            fechas.append(fecha)

    # Seleccionar solo los d√≠as que recibir√°n cifras
    if dias_reciben_cifras > len(fechas):
        dias_reciben_cifras = len(fechas)

    return sorted(random.sample(fechas, dias_reciben_cifras))

def distribuir_cantidad(cantidad_total, fechas_seleccionadas):
    """Distribuye la cantidad total entre las fechas seleccionadas."""
    datos = []
    dias_con_dos_cantidades = random.randint(4, 8) if len(fechas_seleccionadas) < 19 else 0
    dias_una_cantidad = len(fechas_seleccionadas) - dias_con_dos_cantidades
    cantidades_diarias = []

    for _ in range(dias_una_cantidad):
        cantidad = round(random.randint(50, min(350, int(cantidad_total))) / 10) * 10
        cantidades_diarias.append([int(cantidad)])

    for _ in range(dias_con_dos_cantidades):
        cantidad1 = round(random.randint(50, 300) / 10) * 10
        cantidad2 = round(random.randint(50, min(350 - cantidad1, 350)) / 10) * 10
        cantidades_diarias.append([int(cantidad1), int(cantidad2)])

    factura_num = 1
    for i, fecha in enumerate(fechas_seleccionadas):
        facturas = []
        for j, cantidad in enumerate(cantidades_diarias[i]):
            if j == 0:
                facturas.append({
                    'factura_num': factura_num,
                    'cantidad': cantidad,
                    'total': None if len(cantidades_diarias[i]) > 1 else sum(cantidades_diarias[i]),
                    'fecha': fecha.strftime("%d/%m/%Y")
                })
                factura_num += 1
            else:
                facturas.append({
                    'factura_num': None,
                    'cantidad': cantidad,
                    'total': sum(cantidades_diarias[i]),
                    'fecha': fecha.strftime("%d/%m/%Y")
                })
        datos.extend(facturas)

    return datos

def crear_o_actualizar_excel(datos, a√±o, mes, cantidad_total):
    """Crea o actualiza el archivo Excel con los datos proporcionados."""
    nombre_archivo = f"Facturacion_{a√±o}.xlsx"
    mes_nombre = nombre_mes_en_espanol(mes)

    # Obtener el √∫ltimo n√∫mero de factura
    ultimo_num_factura = 0
    if os.path.exists(nombre_archivo):
        wb = load_workbook(nombre_archivo)
        # Buscar el √∫ltimo n√∫mero de factura en todas las hojas anteriores
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Buscar en la columna A el √∫ltimo n√∫mero no vac√≠o
            for cell in ws['A']:
                if cell.value and isinstance(cell.value, int):
                    ultimo_num_factura = max(ultimo_num_factura, cell.value)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = mes_nombre

    # Si la hoja del mes existe, la reemplazamos
    if mes_nombre in wb.sheetnames:
        del wb[mes_nombre]
        ws = wb.create_sheet(mes_nombre)
    elif not os.path.exists(nombre_archivo):
        ws = wb.active
        ws.title = mes_nombre
    else:
        ws = wb.create_sheet(mes_nombre)

    # Configurar encabezados y alineaci√≥n
    ws['A1'] = "FACTURA SIMPLIFICADA"
    ws['B1'] = "CANTIDADES DIARIAS"
    ws['C1'] = "TOTAL"
    ws['D1'] = ""
    ws['E1'] = "FECHA"

    # Alinear la columna A a la izquierda
    alineacion_izquierda = Alignment(horizontal='left')
    ws.column_dimensions['A'].alignment = alineacion_izquierda

    # Actualizar n√∫meros de factura
    fila = 2
    for registro in datos:
        if registro['factura_num'] is not None:
            # Incrementar el n√∫mero de factura continuando la secuencia
            ultimo_num_factura += 1
            celda = ws[f'A{fila}']
            celda.value = ultimo_num_factura
            celda.alignment = alineacion_izquierda  # Alinear cada celda individualmente
        if registro['cantidad'] is not None:
            ws[f'B{fila}'] = f"{registro['cantidad']} ‚Ç¨"
        if registro['total'] is not None:
            ws[f'C{fila}'] = f"{registro['total']} ‚Ç¨"
        if registro['fecha'] is not None:
            ws[f'E{fila}'] = registro['fecha']
        fila += 1

    # Agregar total mensual
    fila += 2
    ws[f'B{fila}'] = f"Total {mes_nombre}"
    ws[f'C{fila}'] = f"{int(cantidad_total)} ‚Ç¨"

    # Ajustar ancho de columnas
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 20

    wb.save(nombre_archivo)
    print(f"\nüëçüèº Datos guardados en {nombre_archivo}, hoja '{mes_nombre}'.")
    return nombre_archivo

def callback(sender, app_data, user_data):                                      # 
    """Callback para manejar el evento del bot√≥n."""
    try:
        a√±o = int(dpg.get_value("a√±o"))
        mes = int(dpg.get_value("mes"))
        cantidad_total = int(dpg.get_value("cantidad_total"))    # cambiar a float
        dias_reciben_cifras = int(dpg.get_value("dias_reciben_cifras"))
        festivos_input = dpg.get_value("festivos")

        # Convertir los d√≠as festivos introducidos en una lista de enteros
        festivos = [int(dia.strip()) for dia in festivos_input.split(",") if dia.strip().isdigit()]

        if mes < 1 or mes > 12:
            raise ValueError("El mes debe estar entre 1 y 12.")
        if cantidad_total <= 0:
            raise ValueError("La cantidad total debe ser mayor que 0.")
        if dias_reciben_cifras < 0:
            raise ValueError("Los d√≠as que reciben cifras no pueden ser negativos.")

        fechas_seleccionadas = generar_fechas_laborables(a√±o, mes, dias_reciben_cifras, len(festivos), festivos)
        datos = distribuir_cantidad(cantidad_total, fechas_seleccionadas)
        archivo = crear_o_actualizar_excel(datos, a√±o, mes, cantidad_total)

        dpg.set_value("output", f"üëçüèº Datos guardados: {archivo}")
    except ValueError as e:
        dpg.set_value("output", f"‚ùå Error: {str(e)}")
    except Exception as e:
        dpg.set_value("output", f"‚ùå Error inesperado: {str(e)}")

# Interfaz gr√°fica con Dear PyGui
dpg.create_context()

with dpg.window(label="Generador de Facturaci√≥n", width=500, height=600):
    dpg.add_input_text(label="A√±o", tag="a√±o", default_value="2025")
    dpg.add_input_text(label="Mes (1-12)", tag="mes", default_value="5")
    dpg.add_input_text(label="Cantidad Total (‚Ç¨)", tag="cantidad_total", default_value="1000")
    dpg.add_input_text(label="D√≠as que reciben cifras", tag="dias_reciben_cifras", default_value="10")
    dpg.add_input_text(label="D√≠as festivos (separados por comas)", tag="festivos", default_value="1, 15")
    dpg.add_button(label="Generar Excel", callback=callback)
    dpg.add_text("", tag="output")

dpg.create_viewport(title='Generador de Facturaci√≥n', width=520, height=640)
dpg.setup_dearpygui()
dpg.show_viewport()
dpg.start_dearpygui()