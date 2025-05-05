import dearpygui.dearpygui as dpg
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from datetime import date
import calendar
import random
import os
import pandas as pd

def nombre_mes_en_espanol(mes):
    """Devuelve el nombre del mes en español."""
    meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
            "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    return meses[mes - 1]

def generar_fechas_laborables(año, mes, dias_reciben_cifras, dias_fiesta, festivos):
    """Genera las fechas laborables del mes, excluyendo domingos y festivos."""
    fechas = []

    # Validar los días festivos
    festivos_validos = []
    _, num_dias = calendar.monthrange(año, mes)
    for dia in festivos:
        if 1 <= dia <= num_dias:
            fecha_festivo = date(año, mes, dia)
            if fecha_festivo.weekday() != 6:  # Excluir domingos
                festivos_validos.append(fecha_festivo)

    # Generar todas las fechas del mes
    for dia in range(1, num_dias + 1):
        fecha = date(año, mes, dia)
        if fecha.weekday() != 6 and fecha not in festivos_validos:
            fechas.append(fecha)

    # Seleccionar solo los días que recibirán cifras
    if dias_reciben_cifras > len(fechas):
        dias_reciben_cifras = len(fechas)

    return sorted(random.sample(fechas, dias_reciben_cifras))

def distribuir_cantidad(cantidad_total, fechas_seleccionadas):
    """Distribuye la cantidad total entre las fechas seleccionadas."""
    datos = []
    total_acumulado = 0
    factura_num = 1
    for fecha in fechas_seleccionadas:
        cantidad = round(random.randint(50, min(350, int(cantidad_total))) / 10) * 10
        total_acumulado += cantidad
        datos.append({
            "Número de Factura": factura_num,
            "Descripción": f"Item {factura_num}",
            "Cantidad": 1,
            "Precio unitario": cantidad,
            "Coste": cantidad,
            "Fecha": fecha.strftime("%d/%m/%Y"),
        })
        factura_num += 1

    datos.append({
        "Número de Factura": None,  # El total no tiene número de factura
        "Descripción": "Total",
        "Cantidad": None,
        "Precio unitario": None,
        "Coste": total_acumulado,  # Aseguramos que total_acumulado sea un número
        "Fecha": None,
    })
    return datos

def crear_o_actualizar_excel(datos, año, mes):
    """Crea o actualiza el archivo Excel con los datos proporcionados."""
    nombre_archivo = f"Facturacion_{año}.xlsx"
    mes_nombre = nombre_mes_en_espanol(mes)

    # Crear el libro de trabajo y la hoja activa
    if os.path.exists(nombre_archivo):
        wb = load_workbook(nombre_archivo)
        if mes_nombre in wb.sheetnames:
            del wb[mes_nombre]  # Elimina la hoja si ya existe
        ws = wb.create_sheet(title=mes_nombre)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = mes_nombre

    # Escribir los encabezados
    ws['A1'] = "FACTURA SIMPLIFICADA"
    ws['A3'] = "Descripción"
    ws['B3'] = "Cantidad"
    ws['C3'] = "Precio unitario"
    ws['D3'] = "Coste"

    # A partir de la fila 4, se escriben los datos
    fila = 4
    for dato in datos:
        ws.cell(row=fila, column=1, value=dato["Descripción"])
        ws.cell(row=fila, column=2, value=dato["Cantidad"])
        precio_unitario = dato.get("Precio unitario")
        coste = dato.get("Coste")
        # Corrección: Convertir a float si no es None y luego formatear
        try:
            ws.cell(row=fila, column=3, value=f"{float(precio_unitario):.2f} €" if precio_unitario is not None else "")
            ws.cell(row=fila, column=4, value=f"{float(coste):.2f} €" if coste is not None else "")
        except ValueError:
            print(f"Error de valor al formatear en la fila {fila}.  Precio unitario: {precio_unitario}, Coste: {coste}")
            ws.cell(row=fila, column=3, value="")
            ws.cell(row=fila, column=4, value="")
        fila += 1

    # Guardar el archivo
    wb.save(nombre_archivo)
    print(f"\n👍🏼 Datos guardados en {nombre_archivo}, hoja '{mes_nombre}'.")
    return nombre_archivo

def callback(sender, app_data, user_data):
    """Callback para manejar el evento del botón."""
    try:
        año = int(dpg.get_value("año"))
        mes = int(dpg.get_value("mes"))
        cantidad_total = int(dpg.get_value("cantidad_total"))
        dias_reciben_cifras = int(dpg.get_value("dias_reciben_cifras"))
        festivos_input = dpg.get_value("festivos")

        # Convertir los días festivos introducidos en una lista de enteros
        festivos = [int(dia.strip()) for dia in festivos_input.split(",") if dia.strip().isdigit()]

        if mes < 1 or mes > 12:
            raise ValueError("El mes debe estar entre 1 y 12.")
        if cantidad_total <= 0:
            raise ValueError("La cantidad total debe ser mayor que 0.")
        if dias_reciben_cifras < 0:
            raise ValueError("Los días que reciben cifras no pueden ser negativos.")

        fechas_seleccionadas = generar_fechas_laborables(año, mes, dias_reciben_cifras, len(festivos), festivos)
        datos = distribuir_cantidad(cantidad_total, fechas_seleccionadas)
        archivo = crear_o_actualizar_excel(datos, año, mes)

        dpg.set_value("output", f"👍🏼 Datos guardados: {archivo}")
        mostrar_popup_extraer_datos() # Llamar a la función para mostrar el popup

    except ValueError as e:
        dpg.set_value("output", f"❌ Error: {str(e)}")
    except Exception as e:
        dpg.set_value("output", f"❌ Error inesperado: {str(e)}")

def extraer_datos_de_fila(sender, app_data, user_data):
    """Extrae los datos de la fila especificada del Excel y genera un nuevo Excel con el formato de la factura."""

    nombre_archivo = dpg.get_value("nombre_archivo_excel")
    fila_a_extraer = int(dpg.get_value("fila_a_extraer"))
    try:
        # Leer el archivo Excel con pandas
        df = pd.read_excel(nombre_archivo)

        # Verificar si la fila existe en el DataFrame
        if fila_a_extraer > len(df):
            dpg.set_value("output_extraer", f"❌ La fila {fila_a_extraer} no existe en el archivo Excel.")
            return

        # Extraer los datos de la fila especificada (las filas en pandas están indexadas desde 0)
        datos_fila = df.iloc[fila_a_extraer - 1].to_dict()  # Restamos 1 porque Excel usa indexación base 1

        # Generar un nuevo Excel con el formato de la factura
        nombre_archivo_salida = f"Factura_Fila_{fila_a_extraer}.xlsx"
        wb_salida = Workbook()
        ws_salida = wb_salida.active

        # Encabezados de la factura
        ws_salida['A1'] = "FACTURA SIMPLIFICADA"
        ws_salida['A3'] = "Descripción"
        ws_salida['B3'] = "Cantidad"
        ws_salida['C3'] = "Precio unitario"
        ws_salida['D3'] = "Coste"

        # Datos de la fila extraída
        ws_salida['A4'] = datos_fila.get('Descripción', '')
        ws_salida['B4'] = datos_fila.get('Cantidad', '')
        precio_unitario = datos_fila.get("Precio unitario")
        coste = datos_fila.get("Coste")
        # Corrección: Convertir a float y luego formatear
        try:
            ws_salida['C4'] = f"{float(precio_unitario):.2f} €" if precio_unitario is not None else ""
            ws_salida['D4'] = f"{float(coste):.2f} €" if coste is not None else ""
        except ValueError:
            print(f"Error de valor al formatear en la extracción de fila. Precio unitario: {precio_unitario}, Coste: {coste}")
            ws_salida['C4'] = ""
            ws_salida['D4'] = ""

        # Total
        ws_salida['C6'] = "Total"
        coste_total = datos_fila.get("Coste") # Obtener el coste
        try:
            ws_salida['D6'] = f"{float(coste_total):.2f} €" if coste_total is not None else "" # Formatear el coste
        except ValueError:
             print(f"Error de valor al formatear el total: {coste_total}")
             ws_salida['D6'] = ""

        wb_salida.save(nombre_archivo_salida)
        dpg.set_value("output_extraer", f"👍🏼 Factura generada en: {nombre_archivo_salida}")
        dpg.configure_item("popup_get_fila", show=False)  # Cerrar la ventana emergente

    except FileNotFoundError:
        dpg.set_value("output_extraer", f"❌ No se encontró el archivo: {nombre_archivo}")
    except Exception as e:
        dpg.set_value("output_extraer", f"❌ Error al leer o generar el archivo Excel: {str(e)}")

def mostrar_popup_extraer_datos():
    """Función para mostrar la ventana emergente de extracción de datos."""
    if not dpg.does_item_exist("popup_get_fila"):  # Verificar si el popup ya existe
        with dpg.window(label="Extraer Datos de Fila", modal=True, tag="popup_get_fila", show=True):
            dpg.add_input_text(label="Nombre del archivo Excel", tag="nombre_archivo_excel")
            dpg.add_input_text(label="Número de fila para extraer datos", tag="fila_a_extraer")
            dpg.add_button(label="Extraer Datos", callback=extraer_datos_de_fila)
            dpg.add_text("", tag="output_extraer")
    else:
        dpg.configure_item("popup_get_fila", show=True) # Si ya existe, solo mostrarlo

# Interfaz gráfica con Dear PyGui
dpg.create_context()

with dpg.window(label="Generador de Facturación", width=800, height=600):
    dpg.add_input_text(label="Año", tag="año", default_value="2025")
    dpg.add_input_text(label="Mes (1-12)", tag="mes", default_value="5")
    dpg.add_input_text(label="Cantidad Total (€)", tag="cantidad_total", default_value="1000")
    dpg.add_input_text(label="Días que reciben cifras", tag="dias_reciben_cifras", default_value="10")
    dpg.add_input_text(label="Días festivos (separados por comas)", tag="festivos", default_value="1, 15")
    dpg.add_button(label="Generar Excel", callback=callback)
    dpg.add_text("", tag="output")

dpg.create_viewport(title='Generador de Facturación', width=800, height=240)
dpg.setup_dearpygui()
dpg.show_viewport()
dpg.start_dearpygui()
