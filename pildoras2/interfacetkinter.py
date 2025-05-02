import tkinter as tk
from tkinter import ttk, messagebox, simpledialog
import openpyxl
from datetime import date, datetime
import calendar
import random
import os

class App(tk.Tk):
    def __init__(self):
        super().__init__()
        self.title("Generador de Facturación")
        self.geometry("500x600")
        self.configure(padx=20, pady=20)

        # Campos de entrada
        self.año = self.crear_campo("Año:")
        self.mes = self.crear_campo("Mes (1-12):")
        self.cantidad_total = self.crear_campo("Cantidad total:")
        self.dias_fiesta = self.crear_campo("Días de fiesta:")
        self.dias_reciben_cifras = self.crear_campo("Días que recibirán cifras:")

        # Botón
        self.btn = ttk.Button(self, text="Generar Excel", command=self.generar_datos)
        self.btn.pack(pady=20)

    def crear_campo(self, label_text):
        frame = ttk.Frame(self)
        frame.pack(fill='x', pady=5)
        ttk.Label(frame, text=label_text).pack(side='left')
        entry = ttk.Entry(frame)
        entry.pack(side='right', expand=True, fill='x')
        return entry

    def generar_datos(self):
        try:
            año = int(self.año.get())
            mes = int(self.mes.get())
            cantidad_total = float(self.cantidad_total.get())
            dias_fiesta = int(self.dias_fiesta.get())
            dias_reciben_cifras = int(self.dias_reciben_cifras.get())

            # Validar los datos
            if mes < 1 or mes > 12:
                raise ValueError("El mes debe estar entre 1 y 12.")
            if cantidad_total <= 0:
                raise ValueError("La cantidad total debe ser mayor que 0.")
            if dias_fiesta < 0 or dias_reciben_cifras < 0:
                raise ValueError("Los días de fiesta y los días que reciben cifras no pueden ser negativos.")

            # Solicitar los días de fiesta
            festivos = []
            for i in range(dias_fiesta):
                dia_festivo = simpledialog.askinteger("Día festivo", f"Introduce el número del día festivo {i+1} (1-31):")
                if dia_festivo is None or dia_festivo < 1 or dia_festivo > 31:
                    raise ValueError("Día festivo inválido.")
                festivos.append(dia_festivo)

            fechas_seleccionadas = self.generar_fechas_laborables(año, mes, dias_reciben_cifras, len(festivos))
            datos = self.distribuir_cantidad(cantidad_total, fechas_seleccionadas)
            self.crear_o_actualizar_excel(datos, año, mes, cantidad_total)

            messagebox.showinfo("Éxito", f"Datos guardados correctamente para {mes}/{año}")
        except ValueError as e:
            messagebox.showerror("Error", str(e))
        except Exception as e:
            messagebox.showerror("Error inesperado", str(e))

    def generar_fechas_laborables(self, año, mes, dias_reciben_cifras, dias_fiesta):
        festivos = []
        for i in range(dias_fiesta):
            dia_festivo = random.randint(1, calendar.monthrange(año, mes)[1])
            fecha_festivo = date(año, mes, dia_festivo)
            festivos.append(fecha_festivo)

        fechas = []
        _, num_dias = calendar.monthrange(año, mes)
        for dia in range(1, num_dias + 1):
            fecha = date(año, mes, dia)
            if fecha.weekday() != 6 and fecha not in festivos:
                fechas.append(fecha)

        if dias_reciben_cifras > len(fechas):
            dias_reciben_cifras = len(fechas)

        return sorted(random.sample(fechas, dias_reciben_cifras))

    def distribuir_cantidad(self, cantidad_total, fechas_seleccionadas):
        datos = []
        dias_con_dos_cantidades = random.randint(4, 8) if len(fechas_seleccionadas) < 19 else 0
        dias_una_cantidad = len(fechas_seleccionadas) - dias_con_dos_cantidades
        cantidades_diarias = []

        for _ in range(dias_una_cantidad):
            cantidad = round(random.randint(50, min(350, int(cantidad_total))) / 10) * 10
            cantidades_diarias.append([int(cantidad)])

        for _ in range(dias_con_dos_cantidades):
            cantidad1 = round(random.randint(50, 300) / 10) * 10
            cantidad2 = round(random.randint(50, min(350 - cantidad1, 350)) / 10) * 10
            cantidades_diarias.append([int(cantidad1), int(cantidad2)])

        factura_num = 1
        for i, fecha in enumerate(fechas_seleccionadas):
            facturas = []
            for j, cantidad in enumerate(cantidades_diarias[i]):
                if j == 0:
                    facturas.append({
                        'factura_num': factura_num,
                        'cantidad': cantidad,
                        'total': None if len(cantidades_diarias[i]) > 1 else sum(cantidades_diarias[i]),
                        'fecha': fecha.strftime("%d/%m/%Y")
                    })
                    factura_num += 1
                else:
                    facturas.append({
                        'factura_num': None,
                        'cantidad': cantidad,
                        'total': sum(cantidades_diarias[i]),
                        'fecha': fecha.strftime("%d/%m/%Y")
                    })
            datos.extend(facturas)

        return datos

    def nombre_mes_en_espanol(self, mes):
        meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
                 "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
        return meses[mes - 1]

    def obtener_ultimo_numero_factura(self, nombre_archivo):
        if not os.path.exists(nombre_archivo):
            return 0
        from openpyxl import load_workbook
        wb = load_workbook(nombre_archivo)
        ultimo_numero = 0
        for hoja in wb.sheetnames:
            ws = wb[hoja]
            for celda in ws['A']:
                if isinstance(celda.value, int):
                    ultimo_numero = max(ultimo_numero, celda.value)
        return ultimo_numero

    def crear_o_actualizar_excel(self, datos, año, mes, cantidad_total):
        nombre_archivo = f"Facturacion_{año}.xlsx"
        mes_nombre = self.nombre_mes_en_espanol(mes)

        if not os.path.exists(nombre_archivo):
            wb = openpyxl.Workbook()
            ws = wb.active
            ws.title = mes_nombre
        else:
            from openpyxl import load_workbook
            wb = load_workbook(nombre_archivo)
            if mes_nombre in wb.sheetnames:
                del wb[mes_nombre]
                ws = wb.create_sheet(mes_nombre)
            else:
                ws = wb.create_sheet(mes_nombre)

        ws['A1'] = "FACTURA SIMPLIFICADA"
        ws['B1'] = "CANTIDADES DIARIAS"
        ws['C1'] = "TOTAL"
        ws['D1'] = ""
        ws['E1'] = "FECHA"

        fila = 2
        for registro in datos:
            if registro['factura_num'] is not None:
                ws[f'A{fila}'] = registro['factura_num']
            if registro['cantidad'] is not None:
                ws[f'B{fila}'] = f"{registro['cantidad']} €"
            if registro['total'] is not None:
                ws[f'C{fila}'] = f"{registro['total']} €"
            if registro['fecha'] is not None:
                ws[f'E{fila}'] = registro['fecha']
            fila += 1

        fila += 2
        ws[f'B{fila}'] = f"Total {mes_nombre}"
        ws[f'C{fila}'] = f"{int(cantidad_total)} €"

        for col in ['A', 'B', 'C', 'D', 'E']:
            ws.column_dimensions[col].width = 20

        wb.save(nombre_archivo)

# Iniciar aplicación
if __name__ == "__main__":
    app = App()
    app.mainloop()