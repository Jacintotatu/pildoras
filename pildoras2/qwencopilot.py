import openpyxl
from openpyxl import Workbook
from datetime import datetime, date
import calendar
import random
import os

def obtener_datos():
    """Solicita los datos al usuario."""
    año = int(input("Introduce el año: "))
    mes = int(input("Introduce el mes (1-12): "))
    cantidad_total = float(input("Introduce la cantidad total a repartir: "))
    dias_fiesta = int(input("Introduce la cantidad de días de fiesta (sin contar domingos): "))
    dias_reciben_cifras = int(input(f"¿Cuántos días del mes {mes}/{año} van a recibir cifras (sin contar domingos)? "))

    return año, mes, cantidad_total, dias_fiesta, dias_reciben_cifras

def es_dia_laborable(fecha):
    """Comprueba si una fecha es día laborable (no domingo)."""
    return fecha.weekday() != 6

def generar_fechas_laborables(año, mes, dias_reciben_cifras, dias_fiesta):
    """Genera las fechas laborables del mes, excluyendo domingos y festivos."""
    fechas = []
    festivos = []

    print(f"Introduce los números de los {dias_fiesta} días de fiesta (solo el número del día):")
    for i in range(dias_fiesta):
        while True:
            try:
                dia_festivo = int(input(f"Día de fiesta {i+1}: "))
                if dia_festivo < 1 or dia_festivo > calendar.monthrange(año, mes)[1]:
                    print(f"El día debe estar entre 1 y {calendar.monthrange(año, mes)[1]}. Vuelve a intentarlo.")
                    continue
                fecha_festivo = date(año, mes, dia_festivo)
                festivos.append(fecha_festivo)
                break
            except ValueError:
                print("Entrada inválida. Por favor introduce un número válido para el día.")

    # Generar todas las fechas del mes
    _, num_dias = calendar.monthrange(año, mes)
    for dia in range(1, num_dias + 1):
        fecha = date(año, mes, dia)
        if es_dia_laborable(fecha) and fecha not in festivos:
            fechas.append(fecha)

    # Seleccionar solo los días que recibirán cifras
    if dias_reciben_cifras > len(fechas):
        print(f"Solo hay {len(fechas)} días laborables disponibles (después de restar domingos y festivos).")
        dias_reciben_cifras = len(fechas)

    fechas_seleccionadas = sorted(random.sample(fechas, dias_reciben_cifras))
    return fechas_seleccionadas

def redondear_multiplo_10(numero):
    """Redondea un número al múltiplo de 10 más cercano."""
    return round(numero / 10) * 10

def distribuir_cantidad(cantidad_total, fechas_seleccionadas):
    """Distribuye la cantidad total entre las fechas seleccionadas."""
    datos = []
    cantidad_restante = cantidad_total
    dias_con_dos_cantidades = 0

    if len(fechas_seleccionadas) < 19:
        dias_con_dos_cantidades = random.randint(4, 8)

    dias_una_cantidad = len(fechas_seleccionadas) - dias_con_dos_cantidades
    cantidades_diarias = []

    # Generar días con una sola cantidad
    for _ in range(dias_una_cantidad):
        cantidad = redondear_multiplo_10(random.randint(50, min(350, int(cantidad_total))))
        cantidades_diarias.append([cantidad])
        cantidad_restante -= cantidad

    # Generar días con dos cantidades
    for _ in range(dias_con_dos_cantidades):
        cantidad1 = redondear_multiplo_10(random.randint(50, 300))
        cantidad2 = redondear_multiplo_10(random.randint(50, min(350, 350 - cantidad1)))
        cantidades_diarias.append([cantidad1, cantidad2])
        cantidad_restante -= (cantidad1 + cantidad2)

    # Ajustar para que sume exactamente la cantidad total
    diferencia = round(cantidad_total - (cantidad_total - cantidad_restante))
    if abs(diferencia) > 0:
        idx = random.randint(0, len(cantidades_diarias) - 1)
        if len(cantidades_diarias[idx]) == 1:
            cantidades_diarias[idx][0] = max(50, min(350, redondear_multiplo_10(cantidades_diarias[idx][0] + diferencia)))
        else:
            ajuste = redondear_multiplo_10(diferencia // 2)
            cantidades_diarias[idx][0] = max(50, min(350, redondear_multiplo_10(cantidades_diarias[idx][0] + ajuste)))
            cantidades_diarias[idx][1] = max(50, min(350, redondear_multiplo_10(cantidades_diarias[idx][1] + ajuste)))

    # Asociar las cantidades con las fechas
    factura_num = 1
    for i, fecha in enumerate(fechas_seleccionadas):
        facturas = []
        for j, cantidad in enumerate(cantidades_diarias[i]):
            if j == 0:
                facturas.append({
                    'factura_num': factura_num,
                    'cantidad': cantidad,
                    'total': None if len(cantidades_diarias[i]) > 1 else sum(cantidades_diarias[i]),
                    'fecha': fecha.strftime("%d/%m/%Y")
                })
                factura_num += 1
            else:
                facturas.append({
                    'factura_num': None,
                    'cantidad': cantidad,
                    'total': sum(cantidades_diarias[i]),
                    'fecha': fecha.strftime("%d/%m/%Y")
                })
        datos.extend(facturas)

    return datos

def nombre_mes_en_espanol(mes):
    """Devuelve el nombre del mes en español."""
    meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
             "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    return meses[mes - 1]

def obtener_ultimo_numero_factura(nombre_archivo):
    """Obtiene el último número de factura del archivo Excel."""
    if not os.path.exists(nombre_archivo):
        return 0

    from openpyxl import load_workbook
    wb = load_workbook(nombre_archivo)
    ultimo_numero = 0

    for hoja in wb.sheetnames:
        ws = wb[hoja]
        for celda in ws['A']:
            if celda.value and isinstance(celda.value, int):
                ultimo_numero = max(ultimo_numero, celda.value)

    return ultimo_numero

def crear_o_actualizar_excel(datos, año, mes, cantidad_total):
    """Crea o actualiza el archivo Excel con los datos proporcionados."""
    nombre_archivo = f"Facturacion_{año}.xlsx"
    mes_nombre = nombre_mes_en_espanol(mes)

    # Obtener el último número de factura
    ultimo_numero_factura = obtener_ultimo_numero_factura(nombre_archivo)

    # Crear libro si no existe
    if not os.path.exists(nombre_archivo):
        wb = Workbook()
        ws = wb.active
        ws.title = mes_nombre
    else:
        from openpyxl import load_workbook
        wb = load_workbook(nombre_archivo)
        if mes_nombre in wb.sheetnames:
            print(f"Ya existe un mes {mes_nombre} en el archivo. ¿Quieres sobrescribirlo?")
            respuesta = input("Escribe 's' para sí o cualquier otra tecla para no: ").lower()
            if respuesta == 's':
                del wb[mes_nombre]
                ws = wb.create_sheet(mes_nombre)
            else:
                mes_nombre += f"_{año}_{mes}"
                ws = wb.create_sheet(mes_nombre)
        else:
            ws = wb.create_sheet(mes_nombre)

    # Escribir encabezados
    ws['A1'] = "FACTURA SIMPLIFICADA"
    ws['B1'] = "CANTIDADES DIARIAS"
    ws['C1'] = "TOTAL"
    ws['D1'] = ""
    ws['E1'] = "FECHA"

    # Escribir datos
    fila = 2
    for registro in datos:
        if registro['factura_num'] is not None:
            ws[f'A{fila}'] = registro['factura_num'] + ultimo_numero_factura
        if registro['cantidad'] is not None:
            # Formatear la cantidad con el símbolo € pero mantenerla como entero
            ws[f'B{fila}'] = f"{registro['cantidad']} €"
        if registro['total'] is not None:
            # Formatear el total con el símbolo € pero mantenerlo como entero
            ws[f'C{fila}'] = f"{registro['total']} €"
        if registro['fecha'] is not None:
            ws[f'E{fila}'] = registro['fecha']
        fila += 1

    # Fila en blanco
    fila += 1

    # Fila de total
    ws[f'B{fila}'] = f"Total {mes_nombre}"
    ws[f'C{fila}'] = f"{int(cantidad_total)} €"

    # Ajustar anchos de columna
    for col in ['A', 'B', 'C', 'D', 'E']:
        ws.column_dimensions[col].width = 20

    wb.save(nombre_archivo)
    print(f"\nDatos guardados en {nombre_archivo}, hoja '{mes_nombre}'.")

def main():
    """Función principal."""
    while True:
        print("\n=== Introducir datos para nuevo mes ===")
        año, mes, cantidad_total, dias_fiesta, dias_reciben_cifras = obtener_datos()

        print(f"\nGenerando fechas laborables para {año}-{mes}...")
        fechas_seleccionadas = generar_fechas_laborables(año, mes, dias_reciben_cifras, dias_fiesta)

        print(f"\nDistribuyendo {cantidad_total}€ entre {len(fechas_seleccionadas)} días laborables...")
        datos = distribuir_cantidad(cantidad_total, fechas_seleccionadas)

        print("\nGenerando archivo Excel...")
        crear_o_actualizar_excel(datos, año, mes, cantidad_total)

        continuar = input("\n¿Quieres introducir otro mes? (s/n): ").lower()
        if continuar != 's':
            break

    print("Programa finalizado.")

if __name__ == "__main__":
    main()