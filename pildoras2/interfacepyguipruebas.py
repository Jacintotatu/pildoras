import dearpygui.dearpygui as dpg
from openpyxl import Workbook, load_workbook
from openpyxl.styles import Alignment
from datetime import date
import calendar
import random
import os 

# Listas de nombres y apellidos
NOMBRES_ESPANOLES = [
    "Antonio", "José", "Manuel", "Francisco", "David", "Juan", "José Antonio", "Daniel", 
    "Carlos", "Jesús", "Alejandro", "Miguel", "José Luis", "Francisco Javier", "Rafael",
    "María", "Carmen", "Ana", "Isabel", "Laura", "Cristina", "Marta", "Elena", "Sara",
    "Paula", "Sandra", "Raquel", "Lucía", "Beatriz", "Alba", "Diego", "Pablo", "Luis",
    "Javier", "Sergio", "Jorge", "Alberto", "Fernando", "Ángel", "Mario", "Marcos",
    "José Manuel", "Miguel Ángel", "Pedro", "Andrés", "Ramón", "Raúl", "Vicente",
    "Sofía", "Patricia", "Nuria", "Silvia", "Rosa", "Marina", "Alicia", "Andrea",
    "Rocío", "Julia", "Inés", "Natalia", "Victoria", "Pilar", "Irene", "Carla",
    "Eduardo", "Rubén", "Víctor", "Roberto", "Jaime", "Salvador", "Ricardo", "Felipe",
    "Teresa", "Dolores", "Mercedes", "Manuela", "Rosario", "Antonia", "Carolina",
    "Josefa", "Gloria", "Begoña", "Yolanda", "Catalina", "Lorena", "Eva", "Esther",
    "Gabriel", "Ignacio", "Jordi", "Gonzalo", "Emilio", "Xavier", "Joan", "Marc"
]

NOMBRES_INGLESES = [
    "John", "William", "James", "George", "Michael", 
    "Elizabeth", "Sarah", "Margaret", "Emma", "Victoria"
]

APELLIDOS_ESPANOLES = [
    "García", "González", "Rodríguez", "Fernández", "López", "Martínez", "Sánchez", 
    "Pérez", "Gómez", "Martín", "Jiménez", "Ruiz", "Hernández", "Díaz", "Moreno",
    "Muñoz", "Álvarez", "Romero", "Alonso", "Gutiérrez", "Navarro", "Torres", 
    "Domínguez", "Vázquez", "Ramos", "Gil", "Ramírez", "Serrano", "Blanco", "Molina",
    "Morales", "Suárez", "Ortega", "Delgado", "Castro", "Ortiz", "Rubio", "Marín",
    "Sanz", "Núñez", "Iglesias", "Medina", "Garrido", "Cortés", "Santos", "Castillo",
    "Lozano", "Guerrero", "Cano", "Prieto", "Méndez", "Cruz", "Calvo", "Gallego",
    "Vidal", "León", "Márquez", "Herrera", "Peña", "Flores", "Cabrera", "Campos",
    "Vega", "Fuentes", "Carrasco", "Díez", "Caballero", "Reyes", "Nieto", "Aguilar",
    "Pascual", "Santana", "Herrero", "Lorenzo", "Montero", "Hidalgo", "Giménez",
    "Ibáñez", "Ferrer", "Durán", "Santiago", "Benítez", "Mora", "Vicente", "Vargas",
    "Arias", "Carmona", "Crespo", "Román", "Pastor", "Soto", "Sáez", "Velasco",
    "Moya", "Soler", "Parra", "Esteban", "Bravo", "Gallardo", "Rojas", "Pardo",
    "Merino", "Franco", "Espinosa", "Izquierdo", "Lara", "Rivas", "Silva", "Rivera",
    "Casado", "Arroyo", "Redondo", "Camacho", "Rey", "Vera", "Otero", "Luque",
    "Gálvez", "Segura", "Heredia", "Luna", "Márquez", "Mendoza", "Abad", "Ferrer",
    "Quintana", "Salazar", "Rincón", "Bernal", "Vila", "Escobar", "Robles", "Santamaría",
    "Palacios", "Benito", "Marcos", "Bautista", "Garrido", "Real", "Soria", "Roldan",
    "Valencia", "Menéndez", "Polo", "Aguirre", "Reina", "Paz", "Salas", "Machado",
    "Rico", "Esteban", "Montes", "Sierra", "Guerra", "Varela", "Miranda", "Guillén",
    "Roldán", "Escudero", "Pacheco", "Zamora", "Jurado", "Mateo", "Galán", "Ribera",
    "Tomás", "Salvador", "Bermejo", "Águilar", "Pereira", "Valle", "Moro", "Rosa",
    "Mesa", "Pozo", "Gracia", "Trujillo", "Morán", "Hurtado", "Montes", "Serra",
    "Rueda", "Plaza", "Vela", "Ayala", "Bueno", "Montes", "Serra", "Rovira", "Costa"
]

APELLIDOS_INGLESES = [
    "Smith", "Jones", "Williams", "Brown", "Taylor",
    "Davies", "Wilson", "Evans", "Johnson", "Roberts"
]

def generar_nombre_aleatorio():
    """Genera un nombre aleatorio con apellidos."""
    # Decidir si será un nombre español (90%) o inglés (10%)
    if random.random() < 0.9:  # 90% probabilidad de nombre español
        nombre = random.choice(NOMBRES_ESPANOLES)
        apellido1 = random.choice(APELLIDOS_ESPANOLES)
        apellido2 = random.choice(APELLIDOS_ESPANOLES)
        return f"{nombre} {apellido1} {apellido2}"
    else:  # 10% probabilidad de nombre inglés
        nombre = random.choice(NOMBRES_INGLESES)
        apellido = random.choice(APELLIDOS_INGLESES)
        return f"{nombre} {apellido}"

def nombre_mes_en_espanol(mes):
    """Devuelve el nombre del mes en español."""
    meses = ["Enero", "Febrero", "Marzo", "Abril", "Mayo", "Junio",
            "Julio", "Agosto", "Septiembre", "Octubre", "Noviembre", "Diciembre"]
    return meses[mes - 1]

def generar_fechas_laborables(año, mes, dias_reciben_cifras, dias_fiesta, festivos):
    """Genera las fechas laborables del mes, excluyendo domingos y festivos."""
    fechas = []

    # Validar los días festivos
    festivos_validos = []
    _, num_dias = calendar.monthrange(año, mes)
    for dia in festivos:
        if 1 <= dia <= num_dias:
            fecha_festivo = date(año, mes, dia)
            if fecha_festivo.weekday() != 6:  # Excluir domingos
                festivos_validos.append(fecha_festivo)

    # Generar todas las fechas del mes
    for dia in range(1, num_dias + 1):
        fecha = date(año, mes, dia)
        if fecha.weekday() != 6 and fecha not in festivos_validos:
            fechas.append(fecha)

    # Seleccionar solo los días que recibirán cifras
    if dias_reciben_cifras > len(fechas):
        dias_reciben_cifras = len(fechas)

    return sorted(random.sample(fechas, dias_reciben_cifras))

def distribuir_cantidad(cantidad_total, fechas_seleccionadas):
    """Distribuye la cantidad total entre las fechas seleccionadas."""
    datos = []
    
    # Determinar cuántos días tendrán dos cifras (entre 4 y 8)
    dias_con_dos_cantidades = random.randint(4, 8)
    
    # Asegurarse de que no exceda el total de días disponibles
    dias_con_dos_cantidades = min(dias_con_dos_cantidades, len(fechas_seleccionadas))
    
    # Seleccionar aleatoriamente qué días tendrán dos cifras
    dias_doble_cifra = random.sample(range(len(fechas_seleccionadas)), dias_con_dos_cantidades)
    
    cantidades_diarias = []
    
    # Generar cantidades para cada día
    for i in range(len(fechas_seleccionadas)):
        if i in dias_doble_cifra:
            # Día con dos cantidades
            cantidad1 = round(random.randint(50, 300) / 10) * 10
            cantidad2 = round(random.randint(50, min(350 - cantidad1, 350)) / 10) * 10
            cantidades_diarias.append([int(cantidad1), int(cantidad2)])
        else:
            # Día con una cantidad
            cantidad = round(random.randint(50, min(350, int(cantidad_total))) / 10) * 10
            cantidades_diarias.append([int(cantidad)])

    # Asociar las cantidades con las fechas
    factura_num = 1
    for i, fecha in enumerate(fechas_seleccionadas):
        facturas = []
        cantidades = cantidades_diarias[i]
        
        if len(cantidades) == 1:
            # Si solo hay una cantidad para este día
            facturas.append({
                'factura_num': factura_num,
                'cantidad': cantidades[0],
                'total': cantidades[0],  # Mostrar total en la única fila
                'fecha': fecha.strftime("%d/%m/%Y")
            })
            factura_num += 1
        else:
            # Si hay dos cantidades para este día
            facturas.append({
                'factura_num': factura_num,
                'cantidad': cantidades[0],
                'total': None,  # Primera fila sin total
                'fecha': fecha.strftime("%d/%m/%Y")
            })
            factura_num += 1
            
            facturas.append({
                'factura_num': factura_num,
                'cantidad': cantidades[1],
                'total': sum(cantidades),  # Total solo en la segunda fila
                'fecha': fecha.strftime("%d/%m/%Y")
            })
            factura_num += 1
            
        datos.extend(facturas)

    return datos

def crear_o_actualizar_excel(datos, año, mes, cantidad_total):
    """Crea o actualiza el archivo Excel con los datos proporcionados."""
    nombre_archivo = f"Facturacion_{año}.xlsx"
    mes_nombre = nombre_mes_en_espanol(mes)

    # Obtener el último número de factura
    ultimo_num_factura = 0
    if os.path.exists(nombre_archivo):
        wb = load_workbook(nombre_archivo)
        # Buscar el último número de factura en todas las hojas anteriores
        for sheet_name in wb.sheetnames:
            ws = wb[sheet_name]
            # Buscar en la columna A el último número no vacío
            for cell in ws['A']:
                if cell.value and isinstance(cell.value, int):
                    ultimo_num_factura = max(ultimo_num_factura, cell.value)
    else:
        wb = Workbook()
        ws = wb.active
        ws.title = mes_nombre

    # Si la hoja del mes existe, la reemplazamos
    if mes_nombre in wb.sheetnames:
        del wb[mes_nombre]
        ws = wb.create_sheet(mes_nombre)
    elif not os.path.exists(nombre_archivo):
        ws = wb.active
        ws.title = mes_nombre
    else:
        ws = wb.create_sheet(mes_nombre)

    # Configurar encabezados y alineación
    ws['A1'] = "FACTURA SIMPLIFICADA"
    ws['B1'] = "CLIENTES"  # Nueva columna
    ws['C1'] = "CANTIDADES DIARIAS"
    ws['D1'] = "TOTAL"
    ws['E1'] = ""
    ws['F1'] = "FECHA"

    # Alinear la columna A a la izquierda
    alineacion_izquierda = Alignment(horizontal='left')
    ws.column_dimensions['A'].alignment = alineacion_izquierda

    # Actualizar números de factura y añadir clientes
    fila = 2
    numero_factura = ultimo_num_factura
    for registro in datos:
        # Asignar SIEMPRE un número correlativo, sin repetir
        numero_factura += 1
        celda = ws[f'A{fila}']
        celda.value = numero_factura
        celda.alignment = alineacion_izquierda
        
        # Añadir nombre de cliente
        ws[f'B{fila}'] = generar_nombre_aleatorio()
        
        if registro['cantidad'] is not None:
            ws[f'C{fila}'] = f"{registro['cantidad']} €"
        if registro['total'] is not None:
            ws[f'D{fila}'] = f"{registro['total']} €"
        if registro['fecha'] is not None:
            ws[f'F{fila}'] = registro['fecha']
        fila += 1

    # Agregar total mensual
    fila += 2
    ws[f'C{fila}'] = f"Total {mes_nombre}"
    ws[f'D{fila}'] = f"{int(cantidad_total)} €"

    # Ajustar ancho de columnas (actualizado para incluir la nueva columna)
    for col in ['A', 'B', 'C', 'D', 'E', 'F']:
        ws.column_dimensions[col].width = 30 if col == 'B' else 20  # Columna B más ancha para los nombres

    wb.save(nombre_archivo)
    print(f"\n👍🏼 Datos guardados en {nombre_archivo}, hoja '{mes_nombre}'.")
    return nombre_archivo

def callback(sender, app_data, user_data):
    """Callback para manejar el evento del botón."""
    try:
        año = int(dpg.get_value("año"))
        mes = int(dpg.get_value("mes"))
        cantidad_total = float(dpg.get_value("cantidad_total"))
        dias_reciben_cifras = int(dpg.get_value("dias_reciben_cifras"))
        festivos_input = dpg.get_value("festivos")

        # Convertir los días festivos introducidos en una lista de enteros
        festivos = [int(dia.strip()) for dia in festivos_input.split(",") if dia.strip().isdigit()]

        if mes < 1 or mes > 12:
            raise ValueError("El mes debe estar entre 1 y 12.")
        if cantidad_total <= 0:
            raise ValueError("La cantidad total debe ser mayor que 0.")
        if dias_reciben_cifras < 0:
            raise ValueError("Los días que reciben cifras no pueden ser negativos.")

        fechas_seleccionadas = generar_fechas_laborables(año, mes, dias_reciben_cifras, len(festivos), festivos)
        datos = distribuir_cantidad(cantidad_total, fechas_seleccionadas)
        archivo = crear_o_actualizar_excel(datos, año, mes, cantidad_total)

        dpg.set_value("output", f"👍🏼 Datos guardados: {archivo}")
    except ValueError as e:
        dpg.set_value("output", f"❌ Error: {str(e)}")
    except Exception as e:
        dpg.set_value("output", f"❌ Error inesperado: {str(e)}")

# Interfaz gráfica con Dear PyGui
dpg.create_context()

with dpg.window(label="Generador de Facturación", width=800, height=600):
    dpg.add_input_text(label="Año", tag="año", default_value="2025")
    dpg.add_input_text(label="Mes (1-12)", tag="mes", default_value="5")
    dpg.add_input_text(label="Cantidad Total (€)", tag="cantidad_total", default_value="1000")
    dpg.add_input_text(label="Días que reciben cifras", tag="dias_reciben_cifras", default_value="10")
    dpg.add_input_text(label="Días festivos (separados por comas)", tag="festivos", default_value="1, 15")
    dpg.add_button(label="Generar Excel", callback=callback)
    dpg.add_text("", tag="output")

dpg.create_viewport(title='Generador de Facturación', width=800, height=240)
dpg.setup_dearpygui()
dpg.show_viewport()
dpg.start_dearpygui()